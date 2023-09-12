from django.shortcuts import render,redirect
from app01.models import Department
from app01.utils.pagination import Pagination

def depart_list(request):
    queryset = Department.objects.all()
    page_obj = Pagination(request,queryset,page_size=20)
    context = {
        "queryset": page_obj.page_queryset,
        "page_string": page_obj.html(),
    }

    return render(request,"depart_list1.html",context)

def depart_add(request):
    if request.method == "GET":
        return render(request,"depart_add1.html")
    title = request.POST.get("title")
    Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    nid = request.GET.get("nid")
    Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

def depart_edit(request,nid):
    if request.method == "GET":
        row_obj = Department.objects.filter(id=nid).first()
        return render(request,"depart_edit1.html",{"row_obj":row_obj})
    title = request.POST.get("title")
    Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


""" 批量新增部门 """
from openpyxl import load_workbook
def upload_mutil(request):
    # 获取文件信息
    file_obj = request.FILES.get("exc")
    print(type(file_obj))
    # 打开excel文件
    wb = load_workbook(file_obj)  # 参数可以是：文件的路径名称 或者文件对象
    sheet = wb.worksheets[0]
    # 循环遍历表格数据
    for row in sheet.iter_rows(min_row=2):
        text =row[0].value
        exists = Department.objects.filter(title=text).exists()
        if not exists:
            Department.objects.create(title=text)
    return redirect("/depart/list/")
